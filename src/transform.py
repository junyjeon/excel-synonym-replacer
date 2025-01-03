import pandas as pd
import openpyxl
from title_generator import create_title_combination
from openpyxl.utils.cell import get_column_letter
import time

def generate_titles(file_path, sheet_name, col_selection, synonym_dict, selected_rows, 
                   version_count, progress_callback, log_callback, model, overwrite=True):
    """유의어 치환으로 새로운 제목 생성"""
    try:
        wb = None
        try:
            if not model:
                raise Exception("모델이 없습니다.")
            
            # DataFrame 복사 및 워크북 로드
            df = model._df.copy()
            wb = openpyxl.load_workbook(file_path)
            ws = wb[sheet_name]
            
            # M열(13번째)부터 상품명 열 처리
            start_col = 12  # M열 (0-based index)
            existing_cols = {}  # 이미 존재하는 상품명 열의 위치
            
            # 1. 기존 상품명 열 위치 찾기
            for ver in range(1, version_count + 1):
                col_name = f'상품명_{ver}'
                if col_name in df.columns:
                    existing_cols[ver] = df.columns.get_loc(col_name)
            
            # 2. M열부터 순차적으로 새 열 추가
            next_col = start_col
            for ver in range(1, version_count + 1):
                col_name = f'상품명_{ver}'
                
                # 이미 존재하는 열이면 건너뛰기
                if ver in existing_cols:
                    next_col = max(next_col, existing_cols[ver] + 1)
                    continue
                
                # 새 열 추가
                while len(df.columns) <= next_col:
                    temp_name = f'Column_{len(df.columns)}'
                    df[temp_name] = ''
                
                # 해당 위치의 열 이름을 상품명_N으로 변경
                df.rename(columns={df.columns[next_col]: col_name}, inplace=True)
                existing_cols[ver] = next_col
                next_col += 1
            
            # 처리할 행 인덱스
            df_indices = selected_rows
            
            if log_callback:
                log_callback(f"처리 시작: 총 {len(df_indices)}개 행")
            
            # 각 선택된 행에 대해 처리
            for i, df_idx in enumerate(df_indices, 1):
                try:
                    current_row = df.iloc[df_idx]
                    titles = []
                    
                    # 각 버전별로 새 제목 생성
                    for version in range(version_count):
                        new_title, _ = create_title_combination(
                            current_row,
                            col_selection,
                            synonym_dict,
                            version
                        )
                        
                        if new_title and not new_title.startswith("ERROR"):
                            titles.append(new_title)
                        else:
                            titles.append(df.iloc[df_idx][f'상품명_{version+1}'])
                    
                    # 모든 버전의 제목 열 업데이트
                    for ver, title in enumerate(titles, 1):
                        col_name = f'상품명_{ver}'
                        if overwrite or not df.iloc[df_idx][col_name]:
                            # DataFrame 업데이트
                            df.loc[df_idx, col_name] = title
                            if model:
                                model.update_cell(df_idx, df.columns.get_loc(col_name), title)
                            
                            # 워크시트 업데이트
                            col_letter = get_column_letter(df.columns.get_loc(col_name) + 1)
                            excel_row = df_idx + 2
                            ws[f"{col_letter}{excel_row}"] = title
                    
                    if progress_callback:
                        progress_callback(i, len(df_indices))
                
                except Exception as row_error:
                    if log_callback:
                        log_callback(f"행 {df_idx} 처리 중 오류: {str(row_error)}")
                    continue
            
            # 변경사항 저장
            wb.save(file_path)
            
        finally:
            if wb:
                wb.close()
        
        if log_callback:
            log_callback("\n완료!")
            
    except Exception as e:
        print(f"\n=== 오류 발생 ===\n{str(e)}")
        raise Exception(f"제목 생성 중 오류 발생: {str(e)}")