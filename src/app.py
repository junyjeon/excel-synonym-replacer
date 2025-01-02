# app.py

import os
import json
import platform
import openpyxl

# 운영체제별 플랫폼 자동 설정
if platform.system() == "Windows":
    os.environ["QT_QPA_PLATFORM"] = "windows"
elif platform.system() == "Linux":
    os.environ["QT_QPA_PLATFORM"] = "xcb"

import pandas as pd
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QPushButton, QCheckBox, QSpinBox, QFileDialog, QTableView, QStatusBar,
    QMessageBox, QInputDialog, QProgressDialog, QHeaderView, QTextEdit
)
from PySide6.QtCore import Qt

from synonyms_manager import load_synonym_dict_from_sheets
from transform import generate_titles
from dataframe_model import DataFrameModel


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("유의어 치환기")
        
        # --- 멤버 변수 초기화 ---
        self._df = pd.DataFrame()
        self.file_path = ""
        self.synonym_dict = {}
        self.current_sheet = ""
        self.settings_file = "settings.json"
        self.last_directory = ""
        
        # 체크박스 매핑(브랜드, 색상, 패턴, 소재, 카테고리)
        self.checkbox_mapping = {}
        
        # 데이터프레임 모델 준비
        self._model = DataFrameModel()
        
        # UI 초기화
        self.initUI()
        
        # 설정 로드
        try:
            self.loadSettings()
        except Exception as e:
            print(f"설정 로드 실패: {str(e)}")

    def initUI(self):
        """UI 구성 요소 초기화 및 레이아웃 설정"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # 상단 레이아웃 (파일 열기)
        top_panel = QHBoxLayout()
        self.label_file = QLabel("불러온 파일: (없음)")
        self.btn_open = QPushButton("파일 열기")
        self.btn_open.clicked.connect(self.open_file_dialog)
        top_panel.addWidget(self.label_file)
        top_panel.addWidget(self.btn_open)
        main_layout.addLayout(top_panel)

        # 체크박스 및 옵션 영역
        opt_layout = QHBoxLayout()

        # 왼쪽: 체크박스 그룹
        category_group = QHBoxLayout()
        self.chk_brand = QCheckBox("브랜드")
        self.chk_color = QCheckBox("색상")
        self.chk_pattern = QCheckBox("패턴")
        self.chk_material = QCheckBox("소재")
        self.chk_category = QCheckBox("카테고리")

        # 체크박스 추가
        for chk in [self.chk_brand, self.chk_color, self.chk_pattern, self.chk_material, self.chk_category]:
            category_group.addWidget(chk)
        opt_layout.addLayout(category_group)

        # 구분선
        opt_layout.addWidget(QLabel(" | "))

        # 오른쪽: 설정 그룹
        settings_group = QHBoxLayout()
        
        # 버전 개수 입력
        version_layout = QHBoxLayout()
        label_version = QLabel("버전:")
        self.spin_version = QSpinBox()
        self.spin_version.setMinimum(1)
        self.spin_version.setMaximum(10)
        self.spin_version.setValue(1)
        self.spin_version.setFixedWidth(50)
        version_layout.addWidget(label_version)
        version_layout.addWidget(self.spin_version)
        settings_group.addLayout(version_layout)

        # 덮어쓰기 체크박스
        self.chk_overwrite = QCheckBox("덮어쓰기")
        self.chk_overwrite.setChecked(True)
        settings_group.addWidget(self.chk_overwrite)

        # 실행 버튼
        self.btn_transform = QPushButton("실행")
        self.btn_transform.clicked.connect(self.transform_data)
        settings_group.addWidget(self.btn_transform)

        # 전체 선택 버튼 추가
        self.btn_select_all = QPushButton("전체 선택")
        self.btn_select_all.clicked.connect(self.select_all_rows)
        settings_group.addWidget(self.btn_select_all)

        opt_layout.addLayout(settings_group)
        main_layout.addLayout(opt_layout)

        # 테이블뷰 안내 레이블
        table_guide = QLabel(
            "※ 행 선택 방법:\n"
            "  - 단일 행 선택: 행 클릭\n"
            "  - 연속 선택: Shift + 클릭\n"
            "  - 개별 선택: Ctrl + 클릭"
        )
        table_guide.setStyleSheet("color: #666666;")
        main_layout.addWidget(table_guide)

        # 테이블뷰
        self.table_view = QTableView()
        self.table_view.setSelectionBehavior(QTableView.SelectRows)
        self.table_view.setSelectionMode(QTableView.ExtendedSelection)
        self.table_view.setEditTriggers(QTableView.NoEditTriggers)
        self.table_view.setModel(self._model)

        # 헤더 설정
        header = self.table_view.horizontalHeader()
        header.setStretchLastSection(True)
        # 기본 열 크기 설정
        for i in range(header.count()):
            if i < 5:
                header.setSectionResizeMode(i, QHeaderView.ResizeToContents)
            else:
                header.setSectionResizeMode(i, QHeaderView.Interactive)
                header.setDefaultSectionSize(100)

        # 행 높이
        self.table_view.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)
        self.table_view.verticalHeader().setDefaultSectionSize(25)

        # 선택 변경 시 이벤트
        self.table_view.selectionModel().selectionChanged.connect(self.on_selection_changed)

        self.table_view.setMinimumHeight(400)
        main_layout.addWidget(self.table_view)

        # 로그 텍스트
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMaximumHeight(100)
        main_layout.addWidget(self.log_text)

        # 상태바
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)

        # 체크박스 매핑(설정 저장/로드 용 키 지정)
        self.checkbox_mapping = {
            'brand': self.chk_brand,
            'color': self.chk_color,
            'pattern': self.chk_pattern,
            'material': self.chk_material,
            'category': self.chk_category
        }

    def open_file_dialog(self):
        """엑셀 파일 불러오기 다이얼로그 열기"""
        dlg = QFileDialog(self)
        dlg.setNameFilters(["Excel Files (*.xlsx *.xls)", "All Files (*)"])
        if dlg.exec():
            selected_file = dlg.selectedFiles()[0]
            if selected_file:
                self.load_excel_file(selected_file)
                try:
                    self.synonym_dict = load_synonym_dict_from_sheets(selected_file)
                    if self.synonym_dict:
                        self.status_bar.showMessage(
                            f"유의어 사전 로드 완료: {len(self.synonym_dict)}개 단어"
                        )
                    else:
                        self.show_message(
                            "경고",
                            "유의어 사전 시트(브랜드,색상,패턴,소재,카테고리)를 찾을 수 없습니다."
                        )
                except Exception as e:
                    self.show_message(
                        "경고",
                        f"유의어 사전 로드 실패: {str(e)}"
                    )

    def load_excel_file(self, path: str):
        """엑셀 파일 로드 및 시트 선택"""
        try:
            wb = None
            wb_values = None
            try:
                # 수식 보존을 위해 data_only=False로 읽기
                wb = openpyxl.load_workbook(path)
                # 수식 결과를 얻기 위해 data_only=True로 다시 읽기
                wb_values = openpyxl.load_workbook(path, data_only=True)
                
                if len(wb.sheetnames) > 1:
                    sheet_name, ok = QInputDialog.getItem(
                        self,
                        "시트 선택",
                        "처리할 시트를 선택하세요:",
                        wb.sheetnames,
                        0,
                        False
                    )
                    if not ok:
                        wb.close()
                        return
                else:
                    sheet_name = wb.sheetnames[0]

                ws = wb[sheet_name]
                
                # 데이터를 직접 읽어서 DataFrame 생성
                data = []
                headers = []
                for cell in ws[1]:
                    headers.append(cell.value)
                
                ws_values = wb_values[sheet_name]
                
                for row in ws_values.iter_rows(min_row=2):
                    row_data = []
                    for cell in row:
                        value = cell.value if cell.value is not None else ''
                        row_data.append(str(value))
                    data.append(row_data)
                
                df = pd.DataFrame(data, columns=headers)

                self._df = df
                self._model.setDataFrame(df)
                self.file_path = path
                self.current_sheet = sheet_name
                self.label_file.setText(f"불러온 파일: {path} ({sheet_name})")
                self.status_bar.showMessage(f"{len(df)} 행 로드 완료")

            finally:
                if wb:
                    wb.close()
                if wb_values:
                    wb_values.close()

        except Exception as e:
            self.show_message("오류", f"파일 열기 실패: {e}")

    def transform_data(self):
        """유의어 치환 실행"""
        try:
            # 기본 검증
            if self._df.empty:
                self.show_message("경고", "먼저 파일을 열어주세요.")
                return
            
            if not self.synonym_dict:
                self.show_message("경고", "유의어 사전이 로드되지 않았습니다.")
                return

            # 선택된 카테고리
            col_selection = []
            if self.chk_brand.isChecked(): 
                col_selection.append("브랜드")
            if self.chk_color.isChecked(): 
                col_selection.append("색상")
            if self.chk_pattern.isChecked(): 
                col_selection.append("패턴")
            if self.chk_material.isChecked(): 
                col_selection.append("소재")
            if self.chk_category.isChecked(): 
                col_selection.append("카테고리")

            if not col_selection:
                self.show_message("경고", "하나 이상의 카테고리를 선택해주세요.")
                return

            # 선택된 행(Excel 기준으로 +2 offset)
            selected_rows = [idx.row() + 2 for idx in self.table_view.selectionModel().selectedRows()]
            if not selected_rows:
                self.show_message("경고", "행을 선택해주세요.")
                return

            # 진행 상태 다이얼로그
            progress = QProgressDialog("처리 중...", "취소", 0, len(selected_rows), self)
            progress.setWindowModality(Qt.WindowModal)
            progress.setWindowFlags(progress.windowFlags() | Qt.WindowStaysOnTopHint)
            progress.move(
                self.x() + (self.width() - progress.width()) // 2,
                self.y() + (self.height() - progress.height()) // 2
            )

            # 로그창 초기화
            self.log_text.clear()

            # 선택된 행 로그
            self.update_log(f"선택된 행: {', '.join(map(str, selected_rows))}")

            # 유의어 치환
            generate_titles(
                self.file_path,
                self.current_sheet,
                col_selection,
                self.synonym_dict,
                selected_rows,
                self.spin_version.value(),
                lambda current, total: progress.setValue(current),
                self.update_log,
                self._model,
                overwrite=self.chk_overwrite.isChecked()
            )

        except Exception as e:
            self.show_message("오류", f"제목 생성 중 오류 발생: {str(e)}")

    def on_selection_changed(self, selected, deselected):
        """테이블 선택 변경 시 상태바에 선택 행 수 표시"""
        rows = self.table_view.selectionModel().selectedRows()
        if rows:
            self.status_bar.showMessage(f"{len(rows)}개 행 선택됨")

    def update_log(self, text: str):
        """로그 텍스트에 메시지를 남김"""
        self.log_text.append(text)
        self.log_text.verticalScrollBar().setValue(
            self.log_text.verticalScrollBar().maximum()
        )

    def loadSettings(self):
        """설정 파일 로드"""
        if not os.path.exists(self.settings_file):
            return
        try:
            with open(self.settings_file, 'r', encoding='utf-8') as f:
                settings = json.load(f)

            # 체크박스 상태 복원
            for key, checkbox in self.checkbox_mapping.items():
                if key in settings:
                    checkbox.setChecked(settings[key])

            # 버전 수 복원
            if 'versions' in settings:
                self.spin_version.setValue(settings['versions'])

            # 최근 파일 경로 복원
            self.last_directory = settings.get('last_directory', '')

        except Exception as e:
            print(f"설정 로드 중 오류 발생: {str(e)}")

    def saveSettings(self):
        """설정 저장"""
        settings = {
            'brand': self.chk_brand.isChecked(),
            'color': self.chk_color.isChecked(),
            'pattern': self.chk_pattern.isChecked(),
            'material': self.chk_material.isChecked(),
            'category': self.chk_category.isChecked(),
            'versions': self.spin_version.value(),
            'last_directory': os.path.dirname(self.file_path) if self.file_path else ''
        }
        try:
            with open(self.settings_file, 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"설정 저장 중 오류 발생: {str(e)}")

    def closeEvent(self, event):
        """프로그램 종료 시 설정 저장"""
        self.saveSettings()
        super().closeEvent(event)

    def select_all_rows(self):
        """모든 행 선택"""
        if not self._df.empty:
            self.table_view.selectAll()

    def show_message(self, title, message):
        msg = QMessageBox(parent=self)
        msg.setWindowTitle(title)
        msg.setText(message)
        msg.setStandardButtons(QMessageBox.Ok)
        msg.setDefaultButton(QMessageBox.Ok)
        
        # 화면 중앙에 표시
        msg.setWindowModality(Qt.ApplicationModal)
        msg.move(
            self.x() + (self.width() - msg.width()) // 2,
            self.y() + (self.height() - msg.height()) // 2
        )
        
        return msg.exec()
