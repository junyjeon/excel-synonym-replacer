#step1_synonym_extract.py
import os
from openai import OpenAI
import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
import re

# .env 파일 불러오기
load_dotenv()

# OpenAI 클라이언트 생성
client = OpenAI()

# OS 환경변수에서 OPENAI_API_KEY 읽어오기
client.api_key = os.getenv("OPENAI_API_KEY")
if not client.api_key:
    raise ValueError("[오류] .env에 'OPENAI_API_KEY'가 없거나 로드되지 않았습니다.")

def get_column_prompt(col_name: str, word: str) -> str:
    """
    열 특성에 맞는 프롬프트 반환
    """
    prompts = {
        "브랜드": f"""'{word}'의 한글-영문 표기를 콤마(,)로 구분해서 작성해주세요.
규칙:
1) 영문 브랜드명이면 한글 발음을, 한글 브랜드명이면 영문 표기를 작성
2) 공식 표기만 사용
3) 대소문자 구분 필요
4) 번호나 설명 없이 단어만 작성

입력 예시 1: ADIDAS
출력 예시 1: 아디다스

입력 예시 2: 아디다스
출력 예시 2: ADIDAS""",

        "색상": f"""'{word}' 색상의 유의어 3개만 콤마(,)로 구분해서 작성해주세요.
규칙:
1) 명도/채도가 다른 색상은 제외 (예: 블루≠스카이블루≠네이비)
2) 정확히 동일한 색상명 우선
3) 영문일 경우 대문자로 작성
4) 되도록 한글 표기로 작성
5) 번호나 설명 없이 단어만 작성

입력 예시: 검정
출력 예시: 블랙, BLACK, 검정색""",

        "패턴": f"""'{word}' 패턴의 정확한 동의어 3개만 콤마(,)로 구분해서 작성해주세요.
규칙:
1) 쇼핑몰에서 쓰이는 패턴 표현
2) 정확히 같은 의미의 패턴 우선
3) 다른 패턴은 제외 (예: 무지≠스트라이프)
4) 번호나 설명 없이 단어만 작성

입력 예시: 체크
출력 예시: 체크무늬, 체크패턴, 격자무늬""",

        "소재": f"""'{word}' 소재의 유의어 3개만 콤마(,)로 구분해서 작성해주세요.
규칙:
1) 쇼핑몰에서 쓰이는 소재 표현
2) 정확히 동일한 소재 우선
3) 없다면 가장 유사한 소재 표현 사용
4) 전혀 다른 소재는 제외
5) 혼방/믹스 등은 별개 소재로 취급
6) 번호나 설명 없이 단어만 작성

입력 예시 1: 폴리에스터
출력 예시 1: 폴리, 폴리에스테르

입력 예시 2: 비스코스
출력 예시 2: 레이온, 인견""",

        "카테고리": f"""'{word}' 카테고리의 유의어 3개만 콤마(,)로 구분해서 작성해주세요.
규칙:
1) 정확히 같은 의류 품목 우선
2) 없다면 가장 유사한 표현 사용
3) 의미가 크게 다른 것은 제외
4) 번호나 설명 없이 단어만 작성

입력 예시 1: 맨투맨
출력 예시 1: 스웨트셔츠, 맨투맨티셔츠, 맨투맨티

입력 예시 2: 셔츠블라우스
출력 예시 2: 블라우스셔츠, 셔츠형블라우스, 블라우스""",
    }
    
    return prompts.get(col_name, "")

def clean_gpt_response(response: str, original_word: str) -> list:
    """
    GPT 응답을 정제하여 유의어 리스트로 변환
    """
    # 번호 패턴 제거 (예: 1., 2), (1), 1))
    response = re.sub(r'^\s*\d+[\.\)]\s*', '', response, flags=re.MULTILINE)
    
    # 설명 텍스트 제거 (예: "다음과 같습니다:", "→" 등)
    response = re.sub(r'.*→\s*', '', response)
    response = re.sub(r'다음과\s*같습니다.*:', '', response)
    
    # 괄호 안의 설명 제거 (예: (Olive), (Military Green))
    response = re.sub(r'\s*\([^)]*\)', '', response)
    
    # 콤마로 분리하고 정제
    words = [w.strip() for w in response.split(',') if w.strip()]
    
    # 원본단어와 같은 경우 제외
    words = [w for w in words if w.lower() != original_word.lower()]
    
    return words

def get_synonyms_from_gpt(original_word: str, col_name: str, num_synonyms=3) -> list:
    """
    GPT에게 'original_word'의 유의어를 num_synonyms개 정도 요청하여 리스트로 반환.
    """
    prompt_text = get_column_prompt(col_name, original_word)

    try:
        response = client.chat.completions.create(
            model="gpt-4o-2024-11-20",
            messages=[{"role": "user", "content": prompt_text}],
        )
        content = response.choices[0].message.content.strip()
        
        # 원본단어를 파라미터로 전달하여 중복 체크
        words = clean_gpt_response(content, original_word)
        
        if len(words) < num_synonyms:
            words += [""] * (num_synonyms - len(words))
        
        return words[:num_synonyms]

    except Exception as e:
        print(f"[오류] GPT 호출 실패: {e}")
        return [""] * num_synonyms

def clean_word(word: str) -> str:
    """
    원본 단어에서 특수문자와 불필요한 공백을 제거
    """
    # 특수문자 제거 또는 변환
    replacements = {
        '^': '',
        '[': '',
        ']': '',
        '&': 'and',
        '/': ' ',
        '  ': ' '  # 중복 공백 제거
    }
    
    result = word
    for old, new in replacements.items():
        result = result.replace(old, new)
    
    return result.strip()

def main(skip_existing=True):
    # 1) 원본 엑셀 파일 열기 (openpyxl 사용)
    input_excel = "제목생성+1.1.xlsx"
    
    # header=None으로 설정하여 모든 행을 데이터로 읽기
    df = pd.read_excel(
        input_excel, 
        sheet_name="11M",
        header=None  # 헤더 행 무시
    )
    
    # 열 이름을 0부터 시작하는 숫자로 설정
    df.columns = range(df.shape[1])
    
    # 열 매핑 설정 (엑셀 열 위치 기준)
    col_mapping = {
        "브랜드": 1,    # B열
        "색상": 2,      # C열
        "패턴": 3,      # D열
        "소재": 4,      # E열
        "카테고리": 5   # F열
    }

    # 2) 추출 대상 열 목록
    target_cols = ["색상","패턴","소재","카테고리"]

    # 3) 기존 엑셀 파일 열기
    try:
        wb = load_workbook(input_excel)
    except:
        wb = Workbook()
        wb.remove(wb.active)

    for col_name in target_cols:
        # 해당하는 열 번호로 데이터 추출
        column_data = df[col_mapping[col_name]]
        
        # 2~6행 선택 (첫 행은 헤더이므로 제외)
        first_five = (
            column_data.iloc[1:10]
            .dropna()
            .astype(str)
            .apply(clean_word)
        )
        
        # 중복 제거
        unique_values = pd.unique(first_five).tolist()

        # 빈 문자열이나 의미 없는 값 제거
        unique_values = [v for v in unique_values if v and v not in ['없음', '-', 'none', 'None']]

        print(f"\n{col_name} 열에서 처리할 단어들: {unique_values}")

        # 시트가 있으면 기존 데이터 확인
        if col_name in wb.sheetnames:
            ws = wb[col_name]
            if skip_existing:
                # 기존 데이터에서 원본단어와 유의어 조합 목록 추출
                existing_data = {}
                for row in range(2, ws.max_row + 1):  # 헤더 제외
                    original = ws.cell(row=row, column=1).value
                    if not original:  # 빈 셀이면 스킵
                        continue
                        
                    original = str(original).strip()
                    
                    # 다른 시트는 유의어1,2,3 체크
                    synonyms = []
                    for col in range(2, 5):  # 2,3,4열 (유의어1,2,3)
                        syn = ws.cell(row=row, column=col).value
                        if syn:  # 빈 셀이 아닌 경우만 추가
                            synonyms.append(str(syn).strip())
                    if synonyms:  # 유의어가 하나라도 있는 경우만 저장
                        existing_data[original] = synonyms
                
                # 기존에 없는 단어만 필터링
                filtered_values = []
                for val in unique_values:
                    val_str = str(val).strip()
                    if val_str not in existing_data:  # 원본단어에 대한 유의어가 없는 경우만 추가
                        filtered_values.append(val_str)
                
                unique_values = filtered_values
                
                if not unique_values:
                    print(f"[스킵] {col_name}: 모든 단어가 이미 처리되어 있습니다.")
                    continue

                # 기존 데이터 유지하고 마지막 행부터 추가
                row_idx = ws.max_row + 1
            else:
                # skip_existing이 False일 때만 기존 데이터 삭제
                ws.delete_rows(1, ws.max_row)
                row_idx = 2  # 첫 번째 행은 헤더용
        else:
            ws = wb.create_sheet(title=col_name)
            row_idx = 2  # 첫 번째 행은 헤더용

        # 헤더는 시트가 새로 생성되거나 덮어쓰기할 때만 추가
        if row_idx == 2:
            ws.cell(row=1, column=1, value="원본")
            for i in range(3):
                ws.cell(row=1, column=(2 + i), value=f"유의어{i+1}")

        # 데이터 추가
        for val in unique_values:
            val_str = val.strip()
            if not val_str:
                continue

            synonyms = get_synonyms_from_gpt(val_str, col_name, num_synonyms=3)

            ws.cell(row=row_idx, column=1, value=val_str)
            for i, syn in enumerate(synonyms):
                ws.cell(row=row_idx, column=(2 + i), value=syn)
            row_idx += 1

    # 4) 원본 파일에 저장
    wb.save(input_excel)
    print(f"완료! '{input_excel}' 파일에 시트가 추가/수정되었습니다.")

if __name__ == "__main__":
    # True: 기존 데이터 있으면 스킵, False: 기존 데이터 덮어쓰기
    main(skip_existing=True)
