# synonyms_manager.py
import openpyxl
from typing import Dict, List, Set

def normalize_text(text: str) -> str:
    """텍스트 정규화"""
    return text.strip().upper()

def load_synonym_dict_from_sheets(excel_path: str, sheet_names=None) -> Dict[str, Dict[str, List[str]]]:
    """유의어 사전 로드"""
    if sheet_names is None:
        sheet_names = ["브랜드", "색상", "패턴", "소재", "카테고리"]
        
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        synonym_dict = {}
        
        for category in sheet_names:
            if category not in wb.sheetnames:
                continue
                
            ws = wb[category]
            synonym_dict[category] = {}
            
            # 데이터가 있는 마지막 행까지만 처리
            max_row = ws.max_row
            while max_row > 1:
                if ws.cell(row=max_row, column=1).value:
                    break
                max_row -= 1
            
            # A열=원본, B~Z열=유의어들
            for row_idx in range(2, max_row + 1):
                orig = ws.cell(row=row_idx, column=1).value
                if not orig:
                    continue
                    
                orig = str(orig).strip()
                seen: Set[str] = set()  # 중복 체크용
                synonyms: List[str] = []
                
                # 콤마로 구분된 데이터 처리
                if ',' in orig:
                    parts = [p.strip() for p in orig.split(',')]
                    if len(parts) > 1:
                        orig = parts[0]
                        for part in parts[1:]:
                            norm_part = normalize_text(part)
                            if norm_part and norm_part not in seen:
                                seen.add(norm_part)
                                synonyms.append(part)
                
                # B열부터 순차적으로 유의어 처리
                for col_idx in range(2, ws.max_column + 1):
                    syn = ws.cell(row=row_idx, column=col_idx).value
                    if syn:
                        syn = str(syn).strip()
                        if col_idx == 5:  # E열인 경우
                            # 콤마로 구분된 데이터를 분리
                            gpt_synonyms = [s.strip() for s in syn.split(',')]
                            for gpt_syn in gpt_synonyms:
                                norm_syn = normalize_text(gpt_syn)
                                if norm_syn and norm_syn not in seen:
                                    seen.add(norm_syn)
                                    synonyms.append(gpt_syn)
                        else:
                            norm_syn = normalize_text(syn)
                            if norm_syn and norm_syn not in seen:
                                seen.add(norm_syn)
                                synonyms.append(syn)
                
                if synonyms:  # 유의어가 있을 때만 추가
                    synonym_dict[category][orig] = synonyms
                    
        return synonym_dict
        
    except Exception as e:
        print(f"유의어 사전 로드 실패: {str(e)}")
        return {}


# def test_synonym_dict():
#     """유의어 사전 테스트"""
#     test_cases = [
#         ("나이키", ["NIKE"]),
#         ("검정", ["블랙", "BLACK"]),
#         ("면", ["코튼", "cotton"])
#     ]
    
#     # 테스트용 사전 생성
#     test_dict = {
#         "브랜드": {"나이키": ["NIKE"]},
#         "색상": {"검정": ["블랙", "BLACK"]},
#         "소재": {"면": ["코튼", "cotton"]}
#     }
    
#     for orig, expected_syns in test_cases:
#         for category, syn_dict in test_dict.items():
#             if orig in syn_dict:
#                 actual_syns = syn_dict[orig]
#                 assert set(actual_syns) == set(expected_syns), \
#                     f"\n원본: {orig}\n기대: {expected_syns}\n실제: {actual_syns}"
