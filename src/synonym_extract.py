#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
from openai import OpenAI
import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

# .env 파일 불러오기
load_dotenv()

# OpenAI 클라이언트 생성
client = OpenAI()

# OS 환경변수에서 OPENAI_API_KEY 읽어오기
client.api_key = os.getenv("OPENAI_API_KEY")
if not client.api_key:
    raise ValueError("[오류] .env에 'OPENAI_API_KEY'가 없거나 로드되지 않았습니다.")

def get_column_prompt(col_name: str, word: str) -> str:
    """
    열 특성에 맞는 프롬프트 반환
    """
    prompts = {
        "브랜드": f"""'{word}' 브랜드의 한글과 영문 실제 브랜드 표기를 알려줘.
규칙:
1) 한글로 입력된 브랜드는 공식 영문 표기
2) 영문으로 입력된 브랜드는 공식 한글 표기

예시:
ADIDAS → ADIDAS, 아디다스
랩 → 랩, LAB
아미 → 아미, ARMY""",

        "색상": f"""'{word}' 색상의 유의어 3개만 콤마(,)로 구분해서 작성해줘
규칙:
1) 명도/채도가 다른 색상은 제외 (예: 블루≠스카이블루≠네이비)
2) 번호나 설명 없이 단어만 작성
3) 출력 형식은 '유의어1,유의어2,유의어3'를 지켜줘

예시)
입력: 검정
출력: 흑색, 블랙, 검정색

입력: 녹색
출력: 초록색, 초록, 잔디색""",

        "패턴": f"""'{word}' 패턴의 유의어 3개만 콤마(,)로 구분해서 작성해줘
규칙:
1) 쇼핑몰에서 쓰이는 패턴 표현
2) 다른 패턴은 제외 (예: 무지≠스트라이프)
3) 번호나 설명 없이 단어만 작성
4) 출력 형식은 '유의어1,유의어2,유의어3'을 꼭 지켜줘

예시:
입력: 체크
출력: 체크무늬, 체크패턴, 격자무늬""",

        "소재": f"""'{word}' 소재의 유의어 3개만 콤마(,)로 구분해서 작성해줘
규칙:
1) 쇼핑몰에서 쓰이는 소재 표현
2) 전혀 다른 소재는 제외
3) 혼방/믹스 등은 별개 소재로 취급
4) 번호나 설명 없이 단어만 작성
5) 출력 형식은 '유의어1,유의어2,유의어3'를 꼭 지켜줘

예시:
입력: 폴리에스터
출력: 폴리, 폴리에스테르

입력: 비스코스
출력: 레이온, 인견""",

        "카테고리": f"""'{word}' 카테고리의 유의어 3개만 콤마(,)로 구분해서 작성해줘
규칙:
1) 쇼핑몰에서 쓰이는 상품 분류명
2) 동일 품목의 다른 표현
3) 번호나 설명 없이 단어만 작성
4) 출력 형식은 '유의어1,유의어2,유의어3'를 꼭 지켜줘

예시:
입력: 맨투맨
출력: 스웨트셔츠, 기모맨투맨, 트레이닝"""
    }
    
    return prompts.get(col_name, "")

def get_synonyms_from_gpt(original_word: str, col_name: str, num_synonyms=3) -> list:
    """
    GPT에게 'original_word'의 유의어를 num_synonyms개 정도 요청하여 리스트로 반환.
    """
    prompt_text = get_column_prompt(col_name, original_word)

    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt_text}],
        )
        content = response.choices[0].message.content.strip()
        synonyms = [s.strip() for s in content.split(",") if s.strip()]
        
        if len(synonyms) < num_synonyms:
            synonyms += [""] * (num_synonyms - len(synonyms))
        
        return synonyms[:num_synonyms]

    except Exception as e:
        print(f"[오류] GPT 호출 실패: {e}")
        return [""] * num_synonyms

def clean_word(word: str) -> str:
    """
    원본 단어에서 특수문자와 불필요한 공백을 제거
    """
    # 특수문자 제거 또는 변환
    replacements = {
        '^': '',
        '[': '',
        ']': '',
        '&': 'and',
        '/': ' ',
        '  ': ' '  # 중복 공백 제거
    }
    
    result = word
    for old, new in replacements.items():
        result = result.replace(old, new)
    
    return result.strip()

def main(skip_existing=True):
    # 1) 원본 엑셀 파일 열기 (openpyxl 사용)
    input_excel = "제목생성+1.1.xlsx"
    
    # header=None으로 설정하여 모든 행을 데이터로 읽기
    df = pd.read_excel(
        input_excel, 
        sheet_name="11M",
        header=None  # 헤더 행 무시
    )
    
    # 열 이름을 0부터 시작하는 숫자로 설정
    df.columns = range(df.shape[1])
    
    # 열 매핑 설정 (엑셀 열 위치 기준)
    col_mapping = {
        "브랜드": 1,    # B열
        "색상": 2,      # C열
        "패턴": 3,      # D열
        "소재": 4,      # E열
        "카테고리": 5   # F열
    }

    # 2) 추출 대상 열 목록
    target_cols = ["브랜드","색상","패턴","소재","카테고리"]

    # 3) 기존 엑셀 파일 열기
    try:
        wb = load_workbook(input_excel)
    except:
        wb = Workbook()
        wb.remove(wb.active)

    for col_name in target_cols:
        # 해당하는 열 번호로 데이터 추출
        column_data = df[col_mapping[col_name]]
        
        # 2~4행만 선택 (첫 행은 헤더이므로 제외)
        first_three = (
            column_data.iloc[1:4]  # 인덱스 1,2,3 (2,3,4행) 선택
            .dropna()
            .astype(str)
            .apply(clean_word)
        )
        
        # 중복 제거
        unique_values = pd.unique(first_three).tolist()

        # 빈 문자열이나 의미 없는 값 제거
        unique_values = [v for v in unique_values if v and v not in ['없음', '-', 'none', 'None']]

        print(f"\n{col_name} 열에서 처리할 단어들: {unique_values}")

        # 시트가 있으면 기존 데이터 확인
        if col_name in wb.sheetnames:
            ws = wb[col_name]
            if skip_existing:
                # 기존 데이터에서 원본단어 목록 추출
                existing_words = set()
                for row in range(2, ws.max_row + 1):  # 헤더 제외
                    # 원본단어와 유의어1,2,3 모두 체크
                    for col in range(1, 5):  # A,B,C,D열
                        value = ws.cell(row=row, column=col).value
                        if value:
                            existing_words.add(str(value).strip())
                
                # 기존에 없는 단어만 필터링
                unique_values = [v for v in unique_values if v not in existing_words]
                
                if not unique_values:
                    print(f"[스킵] {col_name}: 모든 단어가 이미 처리되어 있습니다.")
                    continue
            else:
                # 기존 데이터 삭제
                ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(title=col_name)

        # A열 헤더
        ws.cell(row=1, column=1, value="원본단어")
        for i in range(3):
            ws.cell(row=1, column=(2 + i), value=f"유의어{i+1}")

        row_idx = 2
        for val in unique_values:
            val_str = val.strip()
            if not val_str:
                continue

            synonyms = get_synonyms_from_gpt(val_str, col_name, num_synonyms=3)

            ws.cell(row=row_idx, column=1, value=val_str)
            for i, syn in enumerate(synonyms):
                ws.cell(row=row_idx, column=(2 + i), value=syn)
            row_idx += 1

    # 4) 원본 파일에 저장
    wb.save(input_excel)
    print(f"완료! '{input_excel}' 파일에 시트가 추가/수정되었습니다.")

if __name__ == "__main__":
    # True: 기존 데이터 있으면 스킵, False: 기존 데이터 덮어쓰기
    main(skip_existing=True)
