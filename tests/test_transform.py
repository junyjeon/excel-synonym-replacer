import pandas as pd
import openpyxl
import os
from transform import generate_titles
from unittest.mock import MagicMock

def test_preserve_sheets():
    """다른 시트 보존 테스트"""
    # 1. 여러 시트가 있는 테스트 파일 생성
    test_file = "test_multi_sheets.xlsx"
    
    with pd.ExcelWriter(test_file) as writer:
        # 데이터 시트
        test_df = pd.DataFrame({
            'A': ['', '', ''],
            'B': ['아디다스', 'US폴로아센', '더엣지'],
            'C': ['블랙', '네이비', '스카이블루'],
            'D': ['무지', '스트라이프', '도트무늬'],
            'E': ['면', '폴리', '마혼방'],
            'F': ['맨투맨', '맨투맨티', '니트카라티'],
        }, dtype=str)
        test_df.to_excel(writer, sheet_name='Sheet1', index=False)
        
        # 유의어 사전 시트들
        synonym_sheets = {
            '브랜드': {'나이키': ['NIKE']},
            '색상': {'검정': ['블랙', 'BLACK']},
            '패턴': {'무늬': ['패턴', 'PATTERN']},
            '소재': {'면': ['코튼', 'COTTON']},
            '카테고리': {'상의': ['TOP']}
        }
        
        for sheet_name, data in synonym_sheets.items():
            sheet_df = pd.DataFrame([(k, v[0]) for k, v in data.items()],
                                  columns=['원본', '유의어'])
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    try:
        # Mock 모델 생성
        mock_model = MagicMock()
        mock_model._df = test_df.copy()
        mock_model.update_cell = lambda r,c,v: None
        
        # 데이터 변환 실행
        generate_titles(
            file_path=test_file,
            sheet_name='Sheet1',
            col_selection=['브랜드', '색상'],
            synonym_dict=synonym_sheets,
            selected_rows=[0],
            version_count=3,
            progress_callback=None,
            log_callback=print,
            model=mock_model,  # Mock 모델 사용
            overwrite=True
        )
        
        # 3. 결과 검증
        wb = openpyxl.load_workbook(test_file)
        
        # 모든 시트가 유지되는지 확인
        expected_sheets = {'Sheet1', '브랜드', '색상', '패턴', '소재', '카테고리'}
        assert set(wb.sheetnames) == expected_sheets, \
            f"시트가 유실됨: {expected_sheets - set(wb.sheetnames)}"
        
        # 유의어 시트 내용이 유지되는지 확인
        for sheet_name in ['브랜드', '색상', '패턴', '소재', '카테고리']:
            ws = wb[sheet_name]
            assert ws.cell(row=1, column=1).value == '원본', \
                f"{sheet_name} 시트 내용이 손상됨"
        
        print("모든 시트가 정상적으로 유지됨")
        
    finally:
        if os.path.exists(test_file):
            os.remove(test_file) 

def test_data_transformation():
    """데이터 변환 테스트"""
    test_file = "test_transform.xlsx"
    
    with pd.ExcelWriter(test_file) as writer:
        # 1. 데이터 시트 - 실제 컬럼명 사용
        data_df = pd.DataFrame({
            '유의어': ['', '', ''],
            '브랜드': ['아디다스', '나이키', '푸마'],
            '색상': ['블랙', '화이트', '레드'],
            '패턴': ['무지', '스트라이프', '체크'],
            '소재': ['면', '폴리', '린넨'],
            '카테고리': ['티셔츠', '맨투맨', '셔츠'],
            '상품명_1': ['', '', ''],  # 빈 상품명 컬럼 추가
            '상품명_2': ['', '', '']   # 빈 상품명 컬럼 추가
        }, dtype=str)
        data_df.to_excel(writer, sheet_name='Sheet1', index=False)
        
        # 2. 유의어 시트
        brand_df = pd.DataFrame({
            '원본': ['아디다스', '나이키'],
            '유의어': ['ADIDAS', 'NIKE']
        })
        brand_df.to_excel(writer, sheet_name='브랜드', index=False)
        
        color_df = pd.DataFrame({
            '원본': ['블랙', '화이트'],
            '유의어': ['BLACK', 'WHITE']
        })
        color_df.to_excel(writer, sheet_name='색상', index=False)
    
    try:
        # Mock 모델 생성
        mock_model = MagicMock()
        mock_model._df = data_df.copy()
        mock_model.update_cell = lambda r,c,v: None
        
        # 변환 실행
        generate_titles(
            file_path=test_file,
            sheet_name='Sheet1',
            col_selection=['브랜드', '색상'],
            synonym_dict={
                '브랜드': {'아디다스': ['ADIDAS'], '나이키': ['NIKE']},
                '색상': {'블랙': ['BLACK'], '화이트': ['WHITE']}
            },
            selected_rows=[0, 1],
            version_count=2,
            progress_callback=None,
            log_callback=print,
            model=mock_model,
            overwrite=True
        )
        
        # 4. 결과 검증
        result_df = pd.read_excel(test_file, sheet_name='Sheet1')
        
        # 상품명 컬럼 인덱스 찾기
        col_idx_1 = result_df.columns.get_loc('상품명_1')
        col_idx_2 = result_df.columns.get_loc('상품명_2')
        
        # 4.1 변환된 데이터 확인
        assert result_df.iloc[0, col_idx_1] == 'ADIDAS BLACK 면 티셔츠', "첫 번째 행 변환 실패"
        assert result_df.iloc[1, col_idx_1] == 'NIKE WHITE 스트라이프 맨투맨', "두 번째 행 변환 실패"
        
        # 4.2 원본 데이터 유지 확인
        brand_idx = result_df.columns.get_loc('브랜드')
        color_idx = result_df.columns.get_loc('색상')
        assert result_df.iloc[0, brand_idx] == '아디다스', "원본 브랜드 손상"
        assert result_df.iloc[0, color_idx] == '블랙', "원본 색상 손상"
        
        # 4.3 미처리 행 확인
        third_row_value = result_df.iloc[2, col_idx_1]
        assert pd.isna(third_row_value) or third_row_value.strip() == '', \
            "선택되지 않은 행이 변경됨"
        
        print("데이터 변환 테스트 성공")
        
    finally:
        if os.path.exists(test_file):
            os.remove(test_file)

def test_edge_cases():
    """엣지 케이스 테스트"""
    test_file = "test_edge.xlsx"
    
    try:
        # 1. 빈 데이터프레임
        empty_df = pd.DataFrame(columns=['A', 'B', 'C', 'D', 'E', 'F'])
        empty_df.to_excel(test_file, index=False)
        
        # Mock 모델 생성 (빈 DataFrame)
        mock_model = MagicMock()
        mock_model._df = empty_df.copy()
        mock_model.update_cell = lambda r,c,v: None
        
        generate_titles(
            file_path=test_file,
            sheet_name='Sheet1',
            col_selection=['브랜드', '색상'],
            synonym_dict={},
            selected_rows=[],
            version_count=1,
            progress_callback=None,
            log_callback=print,
            model=mock_model,  # Mock 모델 사용
            overwrite=True
        )
        
        # 2. 특수문자 데이터
        special_df = pd.DataFrame({
            'A': [''],
            'B': ['브랜드!@#'],
            'C': ['색상$%^'],
            'D': ['패턴&*()'],
            'E': ['소재_+'],
            'F': ['카테고리{}']
        })
        special_df.to_excel(test_file, index=False)
        
        # Mock 모델 업데이트 (특수문자 DataFrame)
        mock_model._df = special_df.copy()
        
        generate_titles(
            file_path=test_file,
            sheet_name='Sheet1',
            col_selection=['브랜드', '색상'],
            synonym_dict={},
            selected_rows=[0],
            version_count=1,
            progress_callback=None,
            log_callback=print,
            model=mock_model,  # Mock 모델 사용
            overwrite=True
        )
        
        # 3. 대용량 데이터
        large_df = pd.DataFrame({
            'A': [''] * 1000,
            'B': ['테스트브랜드'] * 1000,
            'C': ['테스트색상'] * 1000,
            'D': ['테스트패턴'] * 1000,
            'E': ['테스트소재'] * 1000,
            'F': ['테스트카테고리'] * 1000
        })
        large_df.to_excel(test_file, index=False)
        
        # Mock 모델 업데이트 (대용량 DataFrame)
        mock_model._df = large_df.copy()
        
        generate_titles(
            file_path=test_file,
            sheet_name='Sheet1',
            col_selection=['브랜드', '색상'],
            synonym_dict={},
            selected_rows=list(range(1000)),
            version_count=1,
            progress_callback=None,
            log_callback=print,
            model=mock_model,  # Mock 모델 사용
            overwrite=True
        )
        
        print("엣지 케이스 테스트 성공")
        
    finally:
        if os.path.exists(test_file):
            os.remove(test_file) 